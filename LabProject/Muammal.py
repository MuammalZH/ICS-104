# Name: Moammal almahfoudh
# ID: 202163190
# Name: Mohammed Al Radhwan
# ID: 202185010


from openpyxl import load_workbook # we only used one function.

wb = load_workbook('sample.xlsx') # to open the workbook that we are going to use.
ws = wb.active # to make the worksheet active.

def main(): # the main function that will handle all the work.     
    while True: # until the user write 8.
        projectMenu()
        userChoice = input("Enter your choice: ") # we called each function for its specified purpose. We made sure that each function has a name
                                                    # that describes its job.
        
        if userChoice == "1":
            dataBase()
                
        elif userChoice == "2":
            findPatient()
                
        elif userChoice == "3":
            newReport()
                
        elif userChoice == "4":
            modifyGrade()
                
        elif userChoice == "5":
            removeReport()
                
        elif userChoice == "6":
            
            saveData_sameFile()

        
        elif userChoice == "7": #saveData_newFile() will return the name of the new file.
            
            nameOfThePath = saveData_newFile()
            wb.save(nameOfThePath) # to save the work.     
                  
        elif userChoice == "8": 
            wb.close() # to close the workbook.
            print("The Application Closed Successfully.") 
            quit() # to exit the main function.
            
            
        else:
            print("You entered an invalid option.") # if the user write anything outside the range 1-8.

    
    
    
def projectMenu(): # This function designs the menu interface.
    numberOfDashes = 55
    print("┌" + "─" * numberOfDashes + "┐") 
    print("│" +"       Breast Imaging-Reporting and Data System        " + "│")
    print("└" + "─" * numberOfDashes + "┘")

    print("┌" + "─" * numberOfDashes + "┐") 
    print("│  1. Read a different dataset from the Excel sheet.    │")
    print("│  2. Find a patient.                                   │")
    print("│  3. Add a new report.                                 │")
    print("│  4. Modify the BI-RADS grade of an existing report.   │")
    print("│  5. Remove an existing report:                        │")
    print("│  6. Save the dataset in the same excel sheet.         │")
    print("│  7. Save the dataset in a new file:                   │")
    print("│  8. Exit                                              │")

    print("└" + "─" * numberOfDashes + "┘")
    print("=" * numberOfDashes) 
    
    
def dataBase():# to Read a different dataset from the Excel sheet.
    DataSetPath = input("Enter The path of the new Excel sheet that contains the dataset: ")
    if DataSetPath.endswith(".xlsx"): # to check if the input ends with .xlsx.
        
        try: # to try opening the file.
            load_workbook(DataSetPath) 
            print("The dataset is found")
            
        except FileNotFoundError:
            print("The dataset was not found")
            main()
    else:
        print("The path is wrong ,check the path and try again")
        

def findPatient():# to find the patient with his data.

    userPID = int(input("The PID of the  patient: "))
    numberOfReports = 0 
    for row in range(2,70):
        excelPID = ws['A' + str(row)].value # the value of the excel index.
        if userPID == excelPID:
            numberOfReports +=1 # to add one every time the user pid is equal to the excel pid

            if numberOfReports == 1: # to print it only once.
                print("PID:", ws['A' + str(row)].value)
                print("Gender:",ws['D' + str(row)].value)
                print("Age:",ws['C' + str(row)].value)
    
    if numberOfReports != 0:
        print("Number of reports:", numberOfReports)
    


    for row in range(2,70):
        excelPID = ws['A' + str(row)].value 
        excelDateOfExam =  ws['E' + str(row)].value
        strExcelDateOfExam = excelDateOfExam.strftime("%d/%m/%Y") # to convert it to string
        if userPID == excelPID:
            print("Date of Exam:",strExcelDateOfExam)
            print("R/L:", ws['B' + str(row)].value)
            print("BI-RADS grade:", ws['F' + str(row)].value)
            


    if numberOfReports == 0:
        print('The patient was not found.')

def newReport(): # to see if  the report of the exact information exists in the dataset or not. If it exists, print this message: “This report exists”; otherwise, add the report to the dataset. Then return to the main menu.
    userPID = int(input("Enter the patient PID: "))
    userRorL = input("Enter the side of the patient breast: ").upper()
    userAge = int(input("Enter the patient age: "))
    userGender = input("Enter patient gender: ").upper()
    userDateOfExam = input("Enter the date of the exam: ")
    userGrade = int(input("Enter BI-RADS Grade: "))

    for row in range(2,70):
        excelPID = ws['A' + str(row)].value
        excelRorL =  ws['B' + str(row)].value
        excelAge =  ws['C' + str(row)].value
        excelGender =  ws['D' + str(row)].value
        excelDateOfExam =  ws['E' + str(row)].value
        strExcelDateOfExam = excelDateOfExam.strftime("%d/%m/%Y")
        excelGrade =  ws['F' + str(row)].value
        
        
        if userPID == excelPID:
            if userRorL == excelRorL:
                if userAge == excelAge:
                    if userGender == excelGender:
                        if userDateOfExam == strExcelDateOfExam:
                            if userGrade == excelGrade:
                                print('This report exists')
                            
                            else:
                                ws.append([userPID,userRorL,userAge,userGender,userDateOfExam,userGrade]) # if the report does not exist, add to the excel file.
                                break        
                        else:
                            ws.append([userPID,userRorL,userAge,userGender,userDateOfExam,userGrade])
                            break
                    else:
                        ws.append([userPID,userRorL,userAge,userGender,userDateOfExam,userGrade])
                        break
                else:
                    ws.append([userPID,userRorL,userAge,userGender,userDateOfExam,userGrade])
                    break
            else:
                ws.append([userPID,userRorL,userAge,userGender,userDateOfExam,userGrade])
                break
        else:
            ws.append([userPID,userRorL,userAge,userGender,userDateOfExam,userGrade])
            break

def modifyGrade(): # . If the report exist, modify the file with the new BI-RADS grade. If the report does not exist, print this message “The report doesn’t exist; choose option 3 to add a new report”. Then return to the main menu.

    userPID = int(input("Enter the patient PID: "))
    userRorL = input("Enter the side of the patient breast: ").upper()
    userAge = int(input("Enter the patient age: "))
    userGender = input("Enter patient gender: ").upper()
    userDateOfExam = input("Enter the date of the exam: ")
    newGrade = 0
    for row in range(2,70):
        excelPID = ws['A' + str(row)].value
        excelRorL =  ws['B' + str(row)].value
        excelAge =  ws['C' + str(row)].value
        excelGender =  ws['D' + str(row)].value
        excelDateOfExam =  ws['E' + str(row)].value
        strExcelDateOfExam = excelDateOfExam.strftime("%d/%m/%Y")
        excelGrade =  ws['F' + str(row)].value
        
        
        if userPID == excelPID:
            if userRorL == excelRorL:
                if userAge == excelAge:
                    if userDateOfExam == strExcelDateOfExam:
                        if userGender == excelGender:
                            newGrade = int(input("Enter BI-RADS Grade: "))
                            ws['F'+ str(row)] = newGrade # to modify the BI-RADS.
                            break
                        else:
                            print("The report doesn't exist; choose option 3 to add a new report")
                            break
                    else:
                        print("The report doesn't exist; choose option 3 to add a new report")
                        break
                else:
                    print("The report doesn't exist; choose option 3 to add a new report")
                    break
            else:
                print("The report doesn't exist; choose option 3 to add a new report")
                break
        else:
            print("The report doesn't exist; choose option 3 to add a new report")
            break

def removeReport(): # if the report exists, delete it from the dataset.  otherwise, print this message: “The report doesn’t exist; please try again” then return to the main menu.

    userPID = int(input("Enter the patient PID: "))
    userRorL = input("Enter the side of the patient breast: ").upper()
    userAge = int(input("Enter the patient age: "))
    userGender = input("Enter patient gender: ").upper()
    userDateOfExam = input("Enter the date of the exam: ")
    userGrade = int(input("Enter BI-RADS Grade: "))

    for row in range(2,70):
        excelPID = ws['A' + str(row)].value
        excelRorL =  ws['B' + str(row)].value
        excelAge =  ws['C' + str(row)].value
        excelDateOfExam =  ws['E' + str(row)].value
        strExcelDateOfExam = excelDateOfExam.strftime("%d/%m/%Y")
        excelGender =  ws['D' + str(row)].value
        excelGrade =  ws['F' + str(row)].value
        
        
        if userPID == excelPID:
            if userRorL == excelRorL:
                if userAge == excelAge:
                    if userGender == excelGender:
                        if userDateOfExam == strExcelDateOfExam:
                            if userGrade == excelGrade:
                                ws.delete_rows(row) # if the report exist , remove the row.
                                print("The row has been deleted successfully.")
                                break
                            else:
                                print("The report doesn't exist; please try again")
                                break        
                        else:
                            print("The report doesn't exist; please try again")
                            break
                    else:
                        print("The report doesn't exist; please try again")
                        break
                else:
                    print("The report doesn't exist; please try again")
                    break
            else:
                print("The report doesn't exist; please try again")
                break
        else:
            print("The report doesn't exist; please try again")
            break
    
def saveData_sameFile(): # This option saves the dataset in the same file.
    wb.save("sample.xlsx")               
    print("The dataset was saved successfully")
    main()
    
def saveData_newFile(): # to Save the dataset in a new file
    filePath = input("The path of the new file ")
    if filePath.endswith(".xlsx"):
        print("The dataset was saved successfully")
        return filePath
    else:
        print("The dataset was not saved successfully; check the path and try again")
        main()
        return False
    

main() # to run whole code.